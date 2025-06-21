from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font
import pandas as pd
import xml.etree.ElementTree as ET
import argparse

# Parse 12-hour time strings like "10:00PM" or "7:00AM"
def parse_time(timestr):
    return datetime.strptime(timestr.strip().upper(), "%I:%M%p").time()

def combine_datetime(date_str, time_str):
    """Combines date and time into the EPG format: YYYYMMDDhhmmss"""
    dt_str = f"{date_str} {time_str}"
    try:
        dt = pd.to_datetime(dt_str)
        return dt.strftime("%Y%m%d%H%M%S")
    except:
        return ""

# Weekly schedule with sample entries
weekly_schedule = {
    "Monday": [
        {"start_time": "7:00AM", "end_time": "9:00AM", "title": "Gospel and Healing Show", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "9:00AM", "end_time": "10:00AM", "title": "Morning Live", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00AM", "end_time": "2:00PM", "title": "AM Show", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "2:00PM", "end_time": "3:00PM", "title": "Amakuru", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "3:00PM", "end_time": "4:00PM", "title": "Movie", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "4:00PM", "end_time": "7:00PM", "title": "Hits Music", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00PM", "end_time": "9:00PM", "title": "Amakuru Yaranze Umuka", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "9:00PM", "end_time": "10:00PM", "title": "Amakuru", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00PM", "end_time": "11:59PM", "title": "Hits Music", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"}
    ],
    "Tuesday": [
        {"start_time": "12:00AM", "end_time": "7:00AM", "title": "Hits Music", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00AM", "end_time": "9:00AM", "title": "Gospel and Healing Show", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "9:00AM", "end_time": "10:00AM", "title": "Morning Live", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00AM", "end_time": "11:00AM", "title": "AM Show", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "11:00AM", "end_time": "12:50PM", "title": "Sports Zone", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "12:50PM", "end_time": "1:50PM", "title": "Amakuru Live", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "1:50PM", "end_time": "3:00PM", "title": "Agasobanuye", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "3:00PM", "end_time": "7:00PM", "title": "News Updates", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00PM", "end_time": "10:00PM", "title": "Amakuru Yaranze Umuka", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00PM", "end_time": "11:59PM", "title": "Hits Music", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"}
    ],
    "Wednesday": [
        {"start_time": "12:00AM", "end_time": "7:00AM", "title": "Hits Music", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00AM", "end_time": "9:00AM", "title": "Gospel and Healing Show", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "9:00AM", "end_time": "10:00AM", "title": "Morning Live", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00AM", "end_time": "2:00PM", "title": "AM Show", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "2:00PM", "end_time": "7:00PM", "title": "Hits Music", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00PM", "end_time": "10:00PM", "title": "Amakuru Yaranze Umuka", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00PM", "end_time": "11:59PM", "title": "Hits Music", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"}
    ],
    "Thursday": [
        {"start_time": "12:00AM", "end_time": "7:00AM", "title": "Hits Music", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00AM", "end_time": "9:00AM", "title": "Gospel and Healing Show", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "9:00AM", "end_time": "10:00AM", "title": "Morning Live", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00AM", "end_time": "2:00PM", "title": "AM Show", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "2:00PM", "end_time": "5:00PM", "title": "Hits Music", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "5:00PM", "end_time": "7:00PM", "title": "The Play Show", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00PM", "end_time": "7:40PM", "title": "Amakuru Yaranze Umuka", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:40PM", "end_time": "9:00PM", "title": "Haramutse Rwanda", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "9:00PM", "end_time": "10:00PM", "title": "Amakuru", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00PM", "end_time": "11:59PM", "title": "Hits Music", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"}
    ],
    "Friday": [
        {"start_time": "12:00AM", "end_time": "7:00AM", "title": "Hits Music", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00AM", "end_time": "9:00AM", "title": "Gospel and Healing Show", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "9:00AM", "end_time": "10:00AM", "title": "Morning Live", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00AM", "end_time": "2:00PM", "title": "AM Show", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "2:00PM", "end_time": "7:00PM", "title": "Hits Music", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00PM", "end_time": "9:00PM", "title": "Amakuru Yaranze Umuka", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "9:00PM", "end_time": "10:00PM", "title": "Amakuru", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00PM", "end_time": "11:59PM", "title": "Hits Music", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"}
    ],
    "Saturday": [
        {"start_time": "12:00AM", "end_time": "7:00AM", "title": "Hits Music", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00AM", "end_time": "8:00PM", "title": "Gospel and Healing Show", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "8:00PM", "end_time": "9:00AM", "title": "Bukombanku", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "9:00AM", "end_time": "10:00AM", "title": "Morning Live", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00AM", "end_time": "2:00PM", "title": "AM Show", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "2:00PM", "end_time": "2:00PM", "title": "Tubimenye", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "2:00PM", "end_time": "7:00PM", "title": "Hits Music", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00PM", "end_time": "7:40PM", "title": "Amakuru Yaranze Umuka", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:40PM", "end_time": "9:00PM", "title": "Haramutse Rwanda", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "9:00PM", "end_time": "10:00PM", "title": "Amakuru", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00PM", "end_time": "11:59PM", "title": "Samedi Détente", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"}
    ],
    "Sunday": [
        {"start_time": "12:00AM", "end_time": "7:00AM", "title": "Hits Music", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00AM", "end_time": "9:00AM", "title": "Gospel and Healing Show", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "9:00AM", "end_time": "10:00AM", "title": "Morning Live", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00AM", "end_time": "2:00PM", "title": "AM Show", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "2:00PM", "end_time": "5:00PM", "title": "Hits Music", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "5:00PM", "end_time": "7:00PM", "title": "Ikiganiro Kidasanzwe", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00PM", "end_time": "7:00PM", "title": "Amakuru Yaranze Umuka", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00PM", "end_time": "9:00PM", "title": "Waramutse Rwanda", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "9:00PM", "end_time": "10:00PM", "title": "Quest Means Business", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00PM", "end_time": "11:00PM", "title": "Amakuru", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "11:00PM", "end_time": "11:59PM", "title": "Hits Music", "icon_url": "https://epg.afromobile.com/storage/uploads/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"}
    ]
}

def generate_epg_excel(start_date_str, weeks=67, output_excel=None):
    """Generate EPG Excel file from schedule data"""
    
    # Set starting Monday date - now in MM/DD/YYYY format
    start_date = datetime.strptime(start_date_str, "%m/%d/%Y")

    # Weekday offset map
    weekday_offsets = {
        "Monday": 0,
        "Tuesday": 1,
        "Wednesday": 2,
        "Thursday": 3,
        "Friday": 4,
        "Saturday": 5,
        "Sunday": 6,
    }

    output_rows = []

    for week in range(weeks):
        for day, programs in weekly_schedule.items():
            base_date = start_date + timedelta(days=weekday_offsets[day]) + timedelta(weeks=week)

            for program in programs:
                stime = parse_time(program["start_time"])
                etime = parse_time(program["end_time"])

                # Determine if the end time is past midnight (i.e., logically next day)
                end_date = base_date
                if etime <= stime:
                    end_date = base_date + timedelta(days=1)

                output_rows.append({
                    "start_date": base_date.strftime("%m/%d/%Y"),  # MM/DD/YYYY format
                    "start_time": program["start_time"],
                    "end_date": end_date.strftime("%m/%d/%Y"),  # MM/DD/YYYY format
                    "end_time": program["end_time"],
                    "channel": "BTN Rwanda",
                    "title": program["title"],
                    "subtitle": "",
                    "description": "",
                    "category": "",
                    "country": "",
                    "episode_number": "",
                    "series_name": "",
                    "actors": "",
                    "director": "",
                    "rating": "",
                    "icon_url": program.get("icon_url", "")
                })

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "EPG Schedule"

    # Write headers
    headers = [
        "start_date", "start_time", "end_date", "end_time", "channel", 
        "title", "subtitle", "description", "category", "country",
        "episode_number", "series_name", "actors", "director", 
        "rating", "icon_url"
    ]

    # Write header row with styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    # Write data rows
    for row_num, row_data in enumerate(output_rows, 2):
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=row_num, column=col_num, value=row_data.get(header, ""))

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Determine output filename if not provided
    if not output_excel:
        output_excel = f"EPG_Schedule_{start_date.strftime('%m_%d_%Y')}_to_{(start_date + timedelta(weeks=weeks)).strftime('%m_%d_%Y')}.xlsx"
    
    # Save the Excel file
    wb.save(output_excel)

    print(f"EPG schedule successfully exported to {output_excel}")
    return output_excel

def excel_to_epg_xml(excel_file, xml_file=None):
    """Convert Excel EPG file to XML format"""
    
    # Read Excel file
    df = pd.read_excel(excel_file)

    # Create <tv> root
    tv = ET.Element("tv")

    # Iterate through each row
    for _, row in df.iterrows():
        # Construct start and stop times
        start = combine_datetime(row["start_date"], row["start_time"])
        stop = combine_datetime(row["end_date"], row["end_time"])

        # Create <programme> with attributes
        programme = ET.SubElement(tv, "programme", {
            "start": start,
            "stop": stop,
            "channel": str(row["channel"])
        })

        # Basic elements
        ET.SubElement(programme, "title").text = str(row["title"])
        ET.SubElement(programme, "sub-title").text = str(row["subtitle"])
        ET.SubElement(programme, "desc").text = str(row["description"])
        ET.SubElement(programme, "category").text = str(row["category"])
        ET.SubElement(programme, "country").text = str(row["country"])
        ET.SubElement(programme, "episode-num").text = str(row["episode_number"])
        ET.SubElement(programme, "series-name").text = str(row["series_name"])
        ET.SubElement(programme, "rating").text = str(row["rating"])
        ET.SubElement(programme, "icon").text = str(row["icon_url"])

        # Credits
        credits = ET.SubElement(programme, "credits")
        ET.SubElement(credits, "actor").text = str(row["actors"])
        ET.SubElement(credits, "director").text = str(row["director"])

    # Determine output filename if not provided
    if not xml_file:
        xml_file = excel_file.replace('.xlsx', '.xml')

    # Write to XML file
    tree = ET.ElementTree(tv)
    tree.write(xml_file, encoding="utf-8", xml_declaration=True)

    print(f"XML file '{xml_file}' generated successfully.")
    return xml_file

def generate_epg(start_date_str, weeks=67, output_excel=None, output_xml=None):
    """Generate both Excel and XML EPG files"""
    
    # First generate the Excel file
    excel_file = generate_epg_excel(start_date_str, weeks, output_excel)
    
    # Then convert to XML
    xml_file = excel_to_epg_xml(excel_file, output_xml)
    
    return excel_file, xml_file

if __name__ == "__main__":
    # Set up command line argument parsing
    parser = argparse.ArgumentParser(description='Generate EPG schedule in Excel and XML formats')
    parser.add_argument('start_date', help='Start date in MM/DD/YYYY format')
    parser.add_argument('--weeks', type=int, default=67, help='Number of weeks to generate')
    parser.add_argument('--excel', help='Output Excel filename (optional)')
    parser.add_argument('--xml', help='Output XML filename (optional)')
    
    args = parser.parse_args()
    
    # Generate both files
    excel_file, xml_file = generate_epg(
        args.start_date,
        weeks=args.weeks,
        output_excel=args.excel,
        output_xml=args.xml
    )
    
    print(f"\nEPG generation complete:")
    print(f"- Excel file: {excel_file}")
    print(f"- XML file: {xml_file}")