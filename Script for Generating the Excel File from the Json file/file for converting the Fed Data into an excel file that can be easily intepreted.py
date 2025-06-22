from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font

# Parse 12-hour time strings like "10:00PM" or "7:00AM"
def parse_time(timestr):
    return datetime.strptime(timestr.strip().upper(), "%I:%M%p").time()

# Weekly schedule with sample entries
weekly_schedule = {
    "Monday": [
        {"start_time": "7:00AM", "end_time": "9:00AM", "title": "Gospel and Healing Show", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "9:00AM", "end_time": "10:00AM", "title": "Morning Live", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00AM", "end_time": "2:00PM", "title": "AM Show", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "2:00PM", "end_time": "3:00PM", "title": "Amakuru", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "3:00PM", "end_time": "4:00PM", "title": "Movie", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "4:00PM", "end_time": "7:00PM", "title": "Hits Music", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00PM", "end_time": "9:00PM", "title": "Amakuru Yaranze Umuka", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00PM", "end_time": "11:59PM", "title": "Hits Music", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"}
    ],
    "Tuesday": [
        {"start_time": "12:00AM", "end_time": "7:00AM", "title": "Hits Music", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00AM", "end_time": "9:00AM", "title": "Gospel and Healing Show", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "9:00AM", "end_time": "10:00AM", "title": "Morning Live", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00AM", "end_time": "11:00AM", "title": "AM Show", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "11:00AM", "end_time": "12:50PM", "title": "Sports Zone", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "12:50PM", "end_time": "1:50PM", "title": "Amakuru Live", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "1:50PM", "end_time": "3:00PM", "title": "Agasobanuye", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "3:00PM", "end_time": "7:00PM", "title": "News Updates", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00PM", "end_time": "10:00PM", "title": "Amakuru Yaranze Umuka", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00PM", "end_time": "11:59PM", "title": "Hits Music", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"}
    ],
    "Wednesday": [
        {"start_time": "12:00AM", "end_time": "7:00AM", "title": "Hits Music", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00AM", "end_time": "9:00AM", "title": "Gospel and Healing Show", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "9:00AM", "end_time": "10:00AM", "title": "Morning Live", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00AM", "end_time": "2:00PM", "title": "AM Show", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "2:00PM", "end_time": "7:00PM", "title": "Hits Music", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00PM", "end_time": "10:00PM", "title": "Amakuru Yaranze Umuka", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00PM", "end_time": "11:59PM", "title": "Hits Music", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"}
    ],
    "Thursday": [
        {"start_time": "12:00AM", "end_time": "7:00AM", "title": "Hits Music", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00AM", "end_time": "9:00AM", "title": "Gospel and Healing Show", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "9:00AM", "end_time": "10:00AM", "title": "Morning Live", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00AM", "end_time": "2:00PM", "title": "AM Show", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "2:00PM", "end_time": "5:00PM", "title": "Hits Music", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "5:00PM", "end_time": "7:00PM", "title": "The Play Show", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00PM", "end_time": "7:40PM", "title": "Amakuru Yaranze Umuka", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:40PM", "end_time": "9:00PM", "title": "Haramutse Rwanda", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "9:00PM", "end_time": "10:00PM", "title": "Amakuru", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00PM", "end_time": "11:59PM", "title": "Hits Music", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"}
    ],
    "Friday": [
        {"start_time": "12:00AM", "end_time": "7:00AM", "title": "Hits Music", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00AM", "end_time": "9:00AM", "title": "Gospel and Healing Show", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "9:00AM", "end_time": "10:00AM", "title": "Morning Live", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00AM", "end_time": "2:00PM", "title": "AM Show", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "2:00PM", "end_time": "7:00PM", "title": "Hits Music", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00PM", "end_time": "9:00PM", "title": "Amakuru Yaranze Umuka", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "9:00PM", "end_time": "10:00PM", "title": "Amakuru", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00PM", "end_time": "11:59PM", "title": "Hits Music", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"}
    ],
    "Saturday": [
        {"start_time": "12:00AM", "end_time": "7:00AM", "title": "Hits Music", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00AM", "end_time": "8:00PM", "title": "Gospel and Healing Show", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "8:00PM", "end_time": "9:00AM", "title": "Bukombanku", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "9:00AM", "end_time": "10:00AM", "title": "Morning Live", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00AM", "end_time": "2:00PM", "title": "AM Show", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "2:00PM", "end_time": "2:00PM", "title": "Tubimenye", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "2:00PM", "end_time": "7:00PM", "title": "Hits Music", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00PM", "end_time": "7:40PM", "title": "Amakuru Yaranze Umuka", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:40PM", "end_time": "9:00PM", "title": "Haramutse Rwanda", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "9:00PM", "end_time": "10:00PM", "title": "Amakuru", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00PM", "end_time": "11:59PM", "title": "Samedi Détente", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"}
    ],
    "Sunday": [
        {"start_time": "12:00AM", "end_time": "7:00AM", "title": "Hits Music", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00AM", "end_time": "9:00AM", "title": "Gospel and Healing Show", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "9:00AM", "end_time": "10:00AM", "title": "Morning Live", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00AM", "end_time": "2:00PM", "title": "AM Show", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "2:00PM", "end_time": "5:00PM", "title": "Hits Music", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "5:00PM", "end_time": "7:00PM", "title": "Ikiganiro Kidasanzwe", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00PM", "end_time": "7:00PM", "title": "Amakuru Yaranze Umuka", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "7:00PM", "end_time": "9:00PM", "title": "Waramutse Rwanda", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "9:00PM", "end_time": "10:00PM", "title": "Quest Means Business", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "10:00PM", "end_time": "11:00PM", "title": "Amakuru", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"},
        {"start_time": "11:00PM", "end_time": "11:59PM", "title": "Hits Music", "icon_url": "https://Example/6xyvD5dWD3svWLatsdUWuyQnSyLy1duz159wcK2f.png"}
    ]
}

# Set starting Monday date - now in MM/DD/YYYY format
start_date = datetime.strptime("6/16/2025", "%m/%d/%Y")

# Weekday offset map
weekday_offsets = {
    "Monday": 0,
    "Tuesday": 1,
    "Wednesday": 2,
    "Thursday": 3,
    "Friday": 4,
    "Saturday": 5,
    "Sunday": 6,
}

weeks = 67 # how many weeks to generate

output_rows = []

for week in range(weeks):
    for day, programs in weekly_schedule.items():
        base_date = start_date + timedelta(days=weekday_offsets[day]) + timedelta(weeks=week)

        for program in programs:
            stime = parse_time(program["start_time"])
            etime = parse_time(program["end_time"])

            # Determine if the end time is past midnight (i.e., logically next day)
            end_date = base_date
            if etime <= stime:
                end_date = base_date + timedelta(days=1)

            output_rows.append({
                "start_date": base_date.strftime("%m/%d/%Y"),  # MM/DD/YYYY format
                "start_time": program["start_time"],
                "end_date": end_date.strftime("%m/%d/%Y"),  # MM/DD/YYYY format
                "end_time": program["end_time"],
                "channel": "BTN Rwanda",
                "title": program["title"],
                "subtitle": "",
                "description": "",
                "category": "",
                "country": "",
                "episode_number": "",
                "series_name": "",
                "actors": "",
                "director": "",
                "rating": "",
                "icon_url": program.get("icon_url", "")
            })

# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "EPG Schedule"

# Write headers
headers = [
    "start_date", "start_time", "end_date", "end_time", "channel", 
    "title", "subtitle", "description", "category", "country",
    "episode_number", "series_name", "actors", "director", 
    "rating", "icon_url"
]

# Write header row with styling
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True)

# Write data rows
for row_num, row_data in enumerate(output_rows, 2):
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=row_num, column=col_num, value=row_data.get(header, ""))

# Auto-adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column_letter].width = adjusted_width

# Save the Excel file
filename = f"EPG_Schedule_{start_date.strftime('%m_%d_%Y')}_to_{(start_date + timedelta(weeks=weeks)).strftime('%m_%d_%Y')}.xlsx"
wb.save(filename)

print(f"EPG schedule successfully exported to {filename}")