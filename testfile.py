import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font
import pandas as pd
import xml.etree.ElementTree as ET
from xml.dom import minidom

def __init__(self, root):
    self.root = root
    self.root.title("EPG Generator")
    self.root.geometry("800x600")
    
    # Initialize variables
    self.start_date = tk.StringVar(value=datetime.now().strftime("%m/%d/%Y"))
    self.weeks = tk.IntVar(value=4)
    self.channel_name = tk.StringVar(value="BTN Rwanda")
    self.output_excel = tk.StringVar()
    self.output_xml = tk.StringVar()
    
    # Initialize schedule data FIRST
    self.weekly_schedule = {
        "Monday": [],
        "Tuesday": [],
        "Wednesday": [],
        "Thursday": [],
        "Friday": [],
        "Saturday": [],
        "Sunday": []
    }
    
    # Load default schedule
    self.load_default_schedule()
    
    # Create main container
    self.main_frame = ttk.Frame(self.root, padding="10")
    self.main_frame.pack(fill=tk.BOTH, expand=True)
    
    # Create widgets
    self.create_input_section()
    self.create_schedule_editor()
    self.create_output_section()
    self.create_buttons()
    
    # Update schedule display
    self.update_schedule_display()
    
    def create_input_section(self):
        input_frame = ttk.LabelFrame(self.main_frame, text="Basic Information", padding="10")
        input_frame.pack(fill=tk.X, pady=5)
        
        # Start Date
        ttk.Label(input_frame, text="Start Date (MM/DD/YYYY):").grid(row=0, column=0, sticky=tk.W)
        ttk.Entry(input_frame, textvariable=self.start_date).grid(row=0, column=1, sticky=tk.W)
        
        # Number of Weeks
        ttk.Label(input_frame, text="Number of Weeks:").grid(row=1, column=0, sticky=tk.W)
        ttk.Spinbox(input_frame, from_=1, to=52, textvariable=self.weeks).grid(row=1, column=1, sticky=tk.W)
        
        # Channel Name
        ttk.Label(input_frame, text="Channel Name:").grid(row=2, column=0, sticky=tk.W)
        ttk.Entry(input_frame, textvariable=self.channel_name).grid(row=2, column=1, sticky=tk.W)
    
    def create_schedule_editor(self):
        schedule_frame = ttk.LabelFrame(self.main_frame, text="Weekly Schedule Editor", padding="10")
        schedule_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Day selector
        self.day_var = tk.StringVar(value="Monday")
        day_selector = ttk.Combobox(schedule_frame, textvariable=self.day_var, 
                                   values=list(self.weekly_schedule.keys()))
        day_selector.grid(row=0, column=0, padx=5, pady=5)
        day_selector.bind("<<ComboboxSelected>>", self.update_schedule_display)
        
        # Schedule entry controls
        entry_frame = ttk.Frame(schedule_frame)
        entry_frame.grid(row=1, column=0, sticky=tk.W, pady=5)
        
        ttk.Label(entry_frame, text="Start Time:").grid(row=0, column=0)
        self.start_time_entry = ttk.Entry(entry_frame, width=10)
        self.start_time_entry.grid(row=0, column=1, padx=5)
        
        ttk.Label(entry_frame, text="End Time:").grid(row=0, column=2)
        self.end_time_entry = ttk.Entry(entry_frame, width=10)
        self.end_time_entry.grid(row=0, column=3, padx=5)
        
        ttk.Label(entry_frame, text="Title:").grid(row=0, column=4)
        self.title_entry = ttk.Entry(entry_frame, width=30)
        self.title_entry.grid(row=0, column=5, padx=5)
        
        ttk.Label(entry_frame, text="Icon URL:").grid(row=0, column=6)
        self.icon_entry = ttk.Entry(entry_frame, width=40)
        self.icon_entry.grid(row=0, column=7, padx=5)
        
        # Buttons for schedule management
        btn_frame = ttk.Frame(schedule_frame)
        btn_frame.grid(row=2, column=0, sticky=tk.W, pady=5)
        
        ttk.Button(btn_frame, text="Add Program", command=self.add_program).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Update Program", command=self.update_program).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Remove Program", command=self.remove_program).pack(side=tk.LEFT, padx=5)
        
        # Schedule display
        self.schedule_tree = ttk.Treeview(schedule_frame, columns=("start", "end", "title", "icon"), 
                                        show="headings", height=10)
        self.schedule_tree.grid(row=3, column=0, sticky=tk.NSEW, pady=5)
        
        self.schedule_tree.heading("start", text="Start Time")
        self.schedule_tree.heading("end", text="End Time")
        self.schedule_tree.heading("title", text="Title")
        self.schedule_tree.heading("icon", text="Icon URL")
        
        self.schedule_tree.column("start", width=100)
        self.schedule_tree.column("end", width=100)
        self.schedule_tree.column("title", width=200)
        self.schedule_tree.column("icon", width=300)
        
        self.schedule_tree.bind("<<TreeviewSelect>>", self.on_program_select)
        
        # Configure grid weights for resizing
        schedule_frame.grid_rowconfigure(3, weight=1)
        schedule_frame.grid_columnconfigure(0, weight=1)
    
    def create_output_section(self):
        output_frame = ttk.LabelFrame(self.main_frame, text="Output Files", padding="10")
        output_frame.pack(fill=tk.X, pady=5)
        
        # Excel File
        ttk.Label(output_frame, text="Excel File:").grid(row=0, column=0, sticky=tk.W)
        ttk.Entry(output_frame, textvariable=self.output_excel).grid(row=0, column=1, sticky=tk.EW)
        ttk.Button(output_frame, text="Browse...", command=lambda: self.browse_file("excel")).grid(row=0, column=2)
        
        # XML File
        ttk.Label(output_frame, text="XML File:").grid(row=1, column=0, sticky=tk.W)
        ttk.Entry(output_frame, textvariable=self.output_xml).grid(row=1, column=1, sticky=tk.EW)
        ttk.Button(output_frame, text="Browse...", command=lambda: self.browse_file("xml")).grid(row=1, column=2)
        
        # Configure grid weights
        output_frame.grid_columnconfigure(1, weight=1)
    
    def create_buttons(self):
        btn_frame = ttk.Frame(self.main_frame)
        btn_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(btn_frame, text="Generate EPG", command=self.generate_epg).pack(side=tk.RIGHT, padx=5)
        ttk.Button(btn_frame, text="Reset", command=self.reset_form).pack(side=tk.RIGHT, padx=5)
        ttk.Button(btn_frame, text="Exit", command=self.root.quit).pack(side=tk.RIGHT, padx=5)
    
    def browse_file(self, file_type):
        if file_type == "excel":
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
                title="Save Excel File As"
            )
            if filename:
                self.output_excel.set(filename)
        else:
            filename = filedialog.asksaveasfilename(
                defaultextension=".xml",
                filetypes=[("XML Files", "*.xml"), ("All Files", "*.*")],
                title="Save XML File As"
            )
            if filename:
                self.output_xml.set(filename)
    
    def load_default_schedule(self):
        # Load your default schedule here
        default_schedule = {
            "Monday": [
                {"start_time": "7:00AM", "end_time": "9:00AM", "title": "Gospel and Healing Show", "icon_url": "https://example.com/icon1.png"},
                {"start_time": "9:00AM", "end_time": "10:00AM", "title": "Morning Live", "icon_url": "https://example.com/icon2.png"}
            ],
            "Tuesday": [
                {"start_time": "7:00AM", "end_time": "9:00AM", "title": "Gospel and Healing Show", "icon_url": "https://example.com/icon1.png"}
            ],
            # Add more days as needed
        }
        
        for day, programs in default_schedule.items():
            self.weekly_schedule[day] = programs.copy()
    
    def update_schedule_display(self, event=None):
        day = self.day_var.get()
        programs = self.weekly_schedule.get(day, [])
        
        # Clear current display
        for item in self.schedule_tree.get_children():
            self.schedule_tree.delete(item)
        
        # Add programs to display
        for program in programs:
            self.schedule_tree.insert("", tk.END, values=(
                program["start_time"],
                program["end_time"],
                program["title"],
                program.get("icon_url", "")
            ))
    
    def on_program_select(self, event):
        selected_item = self.schedule_tree.selection()
        if selected_item:
            values = self.schedule_tree.item(selected_item, "values")
            self.start_time_entry.delete(0, tk.END)
            self.start_time_entry.insert(0, values[0])
            self.end_time_entry.delete(0, tk.END)
            self.end_time_entry.insert(0, values[1])
            self.title_entry.delete(0, tk.END)
            self.title_entry.insert(0, values[2])
            self.icon_entry.delete(0, tk.END)
            self.icon_entry.insert(0, values[3])
    
    def add_program(self):
        day = self.day_var.get()
        start_time = self.start_time_entry.get()
        end_time = self.end_time_entry.get()
        title = self.title_entry.get()
        icon_url = self.icon_entry.get()
        
        if not all([start_time, end_time, title]):
            messagebox.showerror("Error", "Please fill in all required fields")
            return
        
        # Add to schedule
        program = {
            "start_time": start_time,
            "end_time": end_time,
            "title": title,
            "icon_url": icon_url
        }
        
        self.weekly_schedule[day].append(program)
        self.update_schedule_display()
        self.clear_entry_fields()
    
    def update_program(self):
        selected_item = self.schedule_tree.selection()
        if not selected_item:
            messagebox.showerror("Error", "Please select a program to update")
            return
        
        day = self.day_var.get()
        index = self.schedule_tree.index(selected_item[0])
        
        self.weekly_schedule[day][index] = {
            "start_time": self.start_time_entry.get(),
            "end_time": self.end_time_entry.get(),
            "title": self.title_entry.get(),
            "icon_url": self.icon_entry.get()
        }
        
        self.update_schedule_display()
        self.clear_entry_fields()
    
    def remove_program(self):
        selected_item = self.schedule_tree.selection()
        if not selected_item:
            messagebox.showerror("Error", "Please select a program to remove")
            return
        
        day = self.day_var.get()
        index = self.schedule_tree.index(selected_item[0])
        
        del self.weekly_schedule[day][index]
        self.update_schedule_display()
        self.clear_entry_fields()
    
    def clear_entry_fields(self):
        self.start_time_entry.delete(0, tk.END)
        self.end_time_entry.delete(0, tk.END)
        self.title_entry.delete(0, tk.END)
        self.icon_entry.delete(0, tk.END)
    
    def reset_form(self):
        self.start_date.set(datetime.now().strftime("%m/%d/%Y"))
        self.weeks.set(4)
        self.channel_name.set("BTN Rwanda")
        self.output_excel.set("")
        self.output_xml.set("")
        self.load_default_schedule()
        self.update_schedule_display()
        self.clear_entry_fields()
    
    def generate_epg(self):
        try:
            # Generate Excel file
            excel_file = self.generate_epg_excel()
            
            # Generate XML file
            xml_file = self.excel_to_epg_xml(excel_file)
            
            messagebox.showinfo("Success", f"EPG files generated successfully:\nExcel: {excel_file}\nXML: {xml_file}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate EPG files:\n{str(e)}")
    
    def generate_epg_excel(self):
        """Generate EPG Excel file from schedule data"""
        start_date = datetime.strptime(self.start_date.get(), "%m/%d/%Y")
        weeks = self.weeks.get()
        channel_name = self.channel_name.get()
        
        # Weekday offset map
        weekday_offsets = {
            "Monday": 0,
            "Tuesday": 1,
            "Wednesday": 2,
            "Thursday": 3,
            "Friday": 4,
            "Saturday": 5,
            "Sunday": 6,
        }
        
        output_rows = []
        
        for week in range(weeks):
            for day, programs in self.weekly_schedule.items():
                base_date = start_date + timedelta(days=weekday_offsets[day]) + timedelta(weeks=week)
                
                for program in programs:
                    stime = self.parse_time(program["start_time"])
                    etime = self.parse_time(program["end_time"])
                    
                    # Determine if the end time is past midnight (i.e., logically next day)
                    end_date = base_date
                    if etime <= stime:
                        end_date = base_date + timedelta(days=1)
                    
                    output_rows.append({
                        "start_date": base_date.strftime("%m/%d/%Y"),  # MM/DD/YYYY format
                        "start_time": program["start_time"],
                        "end_date": end_date.strftime("%m/%d/%Y"),  # MM/DD/YYYY format
                        "end_time": program["end_time"],
                        "channel": channel_name,
                        "title": program["title"],
                        "subtitle": "",
                        "description": "",
                        "category": "",
                        "country": "",
                        "episode_number": "",
                        "series_name": "",
                        "actors": "",
                        "director": "",
                        "rating": "",
                        "icon_url": program.get("icon_url", "")
                    })
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "EPG Schedule"
        
        # Write headers
        headers = [
            "start_date", "start_time", "end_date", "end_time", "channel", 
            "title", "subtitle", "description", "category", "country",
            "episode_number", "series_name", "actors", "director", 
            "rating", "icon_url"
        ]
        
        # Write header row with styling
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
        
        # Write data rows
        for row_num, row_data in enumerate(output_rows, 2):
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=row_num, column=col_num, value=row_data.get(header, ""))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Determine output filename
        if self.output_excel.get():
            excel_file = self.output_excel.get()
        else:
            excel_file = f"EPG_Schedule_{start_date.strftime('%m_%d_%Y')}_to_{(start_date + timedelta(weeks=weeks)).strftime('%m_%d_%Y')}.xlsx"
        
        # Save the Excel file
        wb.save(excel_file)
        return excel_file
    
    def excel_to_epg_xml(self, excel_file):
        """Convert Excel EPG file to XML format"""
        # Read Excel file
        df = pd.read_excel(excel_file)
        
        # Create <tv> root
        tv = ET.Element("tv")
        
        # Iterate through each row
        for _, row in df.iterrows():
            # Construct start and stop times
            start = self.combine_datetime(row["start_date"], row["start_time"])
            stop = self.combine_datetime(row["end_date"], row["end_time"])
            
            # Create <programme> with attributes
            programme = ET.SubElement(tv, "programme", {
                "start": start,
                "stop": stop,
                "channel": str(row["channel"])
            })
            
            # Only include title and icon as shown in your example
            ET.SubElement(programme, "title").text = str(row["title"])
            if row["icon_url"]:
                ET.SubElement(programme, "icon").text = str(row["icon_url"])
        
        # Determine output filename
        if self.output_xml.get():
            xml_file = self.output_xml.get()
        else:
            xml_file = excel_file.replace('.xlsx', '.xml')
        
        # Write to XML file with pretty formatting
        xmlstr = minidom.parseString(ET.tostring(tv)).toprettyxml(indent="  ")
        
        with open(xml_file, "w", encoding="utf-8") as f:
            f.write(xmlstr)
        
        return xml_file
    
    def parse_time(self, timestr):
        return datetime.strptime(timestr.strip().upper(), "%I:%M%p").time()
    
    def combine_datetime(self, date_str, time_str):
        """Combines date and time into the EPG format: YYYYMMDDhhmmss"""
        dt_str = f"{date_str} {time_str}"
        try:
            dt = pd.to_datetime(dt_str)
            return dt.strftime("%Y%m%d%H%M%S")
        except:
            return ""

if __name__ == "__main__":
    root = tk.Tk()
    app = EPGGeneratorApp(root)
    root.mainloop()