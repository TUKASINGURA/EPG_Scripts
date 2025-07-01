import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font
import pandas as pd
import xml.etree.ElementTree as ET
from xml.dom import minidom
import re

class EPGGeneratorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("©Surmagic IT Solutions")
        
        # Create main container with scrollbar
        self.main_canvas = tk.Canvas(self.root)
        self.main_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Add scrollbar
        self.scrollbar = ttk.Scrollbar(self.root, orient=tk.VERTICAL, command=self.main_canvas.yview)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Configure canvas scrolling
        self.main_canvas.configure(yscrollcommand=self.scrollbar.set)
        self.main_canvas.bind('<Configure>', lambda e: self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all")))
        
        # Create frame inside canvas
        self.main_frame = ttk.Frame(self.main_canvas, padding="10")
        self.main_canvas.create_window((0, 0), window=self.main_frame, anchor="nw")
        # Magenta heading
        heading_label = tk.Label(self.main_frame, text="@Surmagic IT Solutions", font=("Helvetica", 18, "bold"), fg="#FF00FF")
        heading_label.pack(pady=(0, 10))

        # Bind mousewheel for scrolling
        self.main_frame.bind("<MouseWheel>", self._on_mousewheel)
        
        # Initialize variables
        self.start_date = tk.StringVar(value=datetime.now().strftime("%m/%d/%Y"))
        self.weeks = tk.IntVar(value=4)
        self.channel_name = tk.StringVar(value="Chanel Name")
        self.output_excel = tk.StringVar()
        self.output_xml = tk.StringVar()
        
        # Initialize schedule data
        self.weekly_schedule = {
            "Monday": [],
            "Tuesday": [],
            "Wednesday": [],
            "Thursday": [],
            "Friday": [],
            "Saturday": [],
            "Sunday": []
        }
        
        # Create widgets
        self.create_input_section()
        self.create_bulk_entry_section()
        self.create_output_section()
        self.create_action_buttons()
        
        # Make the window resizable
        self.root.minsize(800, 600)
    
    def _on_mousewheel(self, event):
        self.main_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    
    def create_input_section(self):
        input_frame = ttk.LabelFrame(self.main_frame, text="Basic Information", padding="10")
        input_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(input_frame, text="Start Date (MM/DD/YYYY):").grid(row=0, column=0, sticky=tk.W)
        ttk.Entry(input_frame, textvariable=self.start_date).grid(row=0, column=1, sticky=tk.W)
        
        ttk.Label(input_frame, text="Number of Weeks:").grid(row=1, column=0, sticky=tk.W)
        ttk.Spinbox(input_frame, from_=1, to=52, textvariable=self.weeks).grid(row=1, column=1, sticky=tk.W)
        
        ttk.Label(input_frame, text="Channel Name:").grid(row=2, column=0, sticky=tk.W)
        ttk.Entry(input_frame, textvariable=self.channel_name).grid(row=2, column=1, sticky=tk.W)
    
    def create_bulk_entry_section(self):
        bulk_frame = ttk.LabelFrame(self.main_frame, text="Bulk Schedule Entry (Paste tabular data)", padding="10")
        bulk_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Day selector
        self.day_var = tk.StringVar(value="Monday")
        day_selector = ttk.Combobox(bulk_frame, textvariable=self.day_var, 
                                   values=list(self.weekly_schedule.keys()))
        day_selector.grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        day_selector.bind("<<ComboboxSelected>>", self.update_bulk_entry)
        
        # Instructions
        instructions = """Paste your schedule data in tabular format (tab-separated):
start_time\tend_time\ttitle\ticon_url
Example:
6:00am\t10:00am\tBukedde Butya\thttps://example.com/icon1.png
10:00am\t12:00pm\tKalasa Mayanja\thttps://example.com/icon2.png"""
        
        ttk.Label(bulk_frame, text=instructions, wraplength=600).grid(row=0, column=1, sticky=tk.W, padx=10)
        
        # Bulk text entry
        self.bulk_text = scrolledtext.ScrolledText(bulk_frame, width=80, height=15, font=('Consolas', 10))
        self.bulk_text.grid(row=1, column=0, columnspan=2, pady=5, sticky=tk.NSEW)
        
        # Parse button
        ttk.Button(bulk_frame, text="Parse and Save Schedule", command=self.parse_bulk_entry).grid(row=2, column=0, columnspan=2, pady=5)
        
        # Configure grid weights
        bulk_frame.grid_rowconfigure(1, weight=1)
        bulk_frame.grid_columnconfigure(1, weight=1)
    
    def create_output_section(self):
        output_frame = ttk.LabelFrame(self.main_frame, text="Output Files", padding="10")
        output_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(output_frame, text="Excel File:").grid(row=0, column=0, sticky=tk.W)
        ttk.Entry(output_frame, textvariable=self.output_excel, width=40).grid(row=0, column=1, sticky=tk.EW)
        ttk.Button(output_frame, text="Browse...", command=lambda: self.browse_file("excel")).grid(row=0, column=2)
        
        ttk.Label(output_frame, text="XML File:").grid(row=1, column=0, sticky=tk.W)
        ttk.Entry(output_frame, textvariable=self.output_xml, width=40).grid(row=1, column=1, sticky=tk.EW)
        ttk.Button(output_frame, text="Browse...", command=lambda: self.browse_file("xml")).grid(row=1, column=2)
        
        output_frame.grid_columnconfigure(1, weight=1)
    
    def create_action_buttons(self):
        action_frame = ttk.Frame(self.main_frame)
        action_frame.pack(fill=tk.X, pady=10)
        
        # Big green generate button
        self.generate_btn = ttk.Button(
            action_frame, 
            text="Generate EPG Files", 
            command=self.generate_epg,
            style="Big.TButton"
        )
        self.generate_btn.pack(side=tk.RIGHT, padx=5, ipadx=10, ipady=5)
        
        # Style configuration for the big button
        style = ttk.Style()
        style.configure("Big.TButton", font=('Helvetica', 12, 'bold'), foreground='green')
        
        ttk.Button(action_frame, text="Reset Form", command=self.reset_form).pack(side=tk.RIGHT, padx=5)
        ttk.Button(action_frame, text="Exit", command=self.root.quit).pack(side=tk.RIGHT, padx=5)
    
    def update_bulk_entry(self, event=None):
        day = self.day_var.get()
        programs = self.weekly_schedule.get(day, [])
        
        # Clear the text widget
        self.bulk_text.delete(1.0, tk.END)
        
        # Add programs to text widget in tabular format
        for program in programs:
            line = f"{program['start_time']}\t{program['end_time']}\t{program['title']}"
            if program.get('icon_url'):
                line += f"\t{program['icon_url']}"
            self.bulk_text.insert(tk.END, line + "\n")
    
    def parse_bulk_entry(self):
        day = self.day_var.get()
        text = self.bulk_text.get(1.0, tk.END)
        
        # Clear existing programs for this day
        self.weekly_schedule[day] = []
        
        # Parse each line
        for line in text.strip().split('\n'):
            line = line.strip()
            if not line:
                continue

            # Split using tabs or multiple spaces
            parts = re.split(r'\t|\s{2,}', line)
            parts = [part.strip() for part in parts if part.strip()]
            
            if len(parts) < 3:
                continue  # Must have at least start, end, title
            
            # Initialize all fields with empty strings
            program = {
                "start_time": parts[0],
                "end_time": parts[1],
                "title": "",
                "description": "",
                "icon_url": ""
            }
            
            # Find and remove icon URL if present
            for i in reversed(range(len(parts))):
                if parts[i].startswith("http"):
                    program["icon_url"] = parts[i]
                    parts.pop(i)
                    break
            
            # The remaining parts after start_time, end_time, and icon_url removal
            remaining_parts = parts[2:]
            
            # First remaining part is title (even if empty)
            if len(remaining_parts) > 0:
                program["title"] = remaining_parts[0] if remaining_parts[0] else ""
            
            # Second remaining part (if exists) is description
            if len(remaining_parts) > 1:
                program["description"] = remaining_parts[1]

            self.weekly_schedule[day].append(program)
        
        messagebox.showinfo("Success", f"Schedule for {day} saved with {len(self.weekly_schedule[day])} programs!")
    
    def browse_file(self, file_type):
        if file_type == "excel":
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files", ".xlsx"), ("All Files", ".*")],
                title="Save Excel File As"
            )
            if filename:
                self.output_excel.set(filename)
        else:
            filename = filedialog.asksaveasfilename(
                defaultextension=".xml",
                filetypes=[("XML Files", ".xml"), ("All Files", ".*")],
                title="Save XML File As"
            )
            if filename:
                self.output_xml.set(filename)
    
    def reset_form(self):
        self.start_date.set(datetime.now().strftime("%m/%d/%Y"))
        self.weeks.set(4)
        self.channel_name.set("BTN Rwanda")
        self.output_excel.set("")
        self.output_xml.set("")
        self.weekly_schedule = {day: [] for day in self.weekly_schedule}
        self.update_bulk_entry()
    
    def generate_epg(self):
        try:
            # Validate inputs
            missing_days = [day for day, programs in self.weekly_schedule.items() if len(programs) == 0]
            if missing_days:
                missing_list = ", ".join(missing_days)
                raise ValueError(f"Please enter schedule data for: {missing_list}")

            if not self.start_date.get():
                raise ValueError("Please specify a start date")
                
            excel_file = self.generate_epg_excel()
            xml_file = self.excel_to_epg_xml(excel_file)
            
            messagebox.showinfo(
                "Success", 
                f"EPG files generated successfully!\n\n"
                f"Excel file: {excel_file}\n"
                f"XML file: {xml_file}",
                icon='info'
            )
        except Exception as e:
            messagebox.showerror(
                "Generation Error",
                f"Failed to generate EPG files:\n\n{str(e)}",
                icon='error'
            )
    
    def generate_epg_excel(self):
        start_date = datetime.strptime(self.start_date.get(), "%m/%d/%Y")
        weeks = self.weeks.get()
        channel_name = self.channel_name.get()
        
        weekday_offsets = {
            "Monday": 0,
            "Tuesday": 1,
            "Wednesday": 2,
            "Thursday": 3,
            "Friday": 4,
            "Saturday": 5,
            "Sunday": 6,
        }
        
        output_rows = []
        
        for week in range(weeks):
            for day, programs in self.weekly_schedule.items():
                base_date = start_date + timedelta(days=weekday_offsets[day]) + timedelta(weeks=week)
                
                for program in programs:
                    stime = self.parse_time(program["start_time"])
                    etime = self.parse_time(program["end_time"])
                    
                    end_date = base_date
                    if etime <= stime:
                        end_date = base_date + timedelta(days=1)
                    
                    output_rows.append({
                        "start_date": base_date.strftime("%m/%d/%Y"),
                        "start_time": program["start_time"],
                        "end_date": end_date.strftime("%m/%d/%Y"),
                        "end_time": program["end_time"],
                        "channel": channel_name,
                        "title": program.get("title", "") or "",
                        "subtitle": "",
                        "description": program.get("description", ""),
                        "category": "",
                        "country": "",
                        "episode_number": "",
                        "series_name": "",
                        "actors": "",
                        "director": "",
                        "rating": "",
                        "icon_url": program.get("icon_url", "")
                    })
        
        wb = Workbook()
        ws = wb.active
        ws.title = "EPG Schedule"
        
        headers = [
            "start_date", "start_time", "end_date", "end_time", "channel", 
            "title", "subtitle", "description", "category", "country",
            "episode_number", "series_name", "actors", "director", 
            "rating", "icon_url"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
        
        for row_num, row_data in enumerate(output_rows, 2):
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=row_num, column=col_num, value=row_data.get(header, ""))
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        if self.output_excel.get():
            excel_file = self.output_excel.get()
        else:
            excel_file = f"EPG_Schedule_{start_date.strftime('%m_%d_%Y')}to{(start_date + timedelta(weeks=weeks)).strftime('%m_%d_%Y')}.xlsx"
        
        wb.save(excel_file)
        return excel_file
    
    def excel_to_epg_xml(self, excel_file):
        df = pd.read_excel(excel_file)
        tv = ET.Element("tv")
        
        for _, row in df.iterrows():
            start = self.combine_datetime(row["start_date"], row["start_time"])
            stop = self.combine_datetime(row["end_date"], row["end_time"])
            
            # # Skip entries with empty titles
            # if you need all rows from the Excel file tot  included in the XML,
            #  even if the title or description is missing, and you do not want a placeholder
            # if pd.isna(row["title"]) or not str(row["title"]).strip():
            #     continue
            
            programme = ET.SubElement(tv, "programme", {
                "start": start,
                "stop": stop,
            })
            ET.SubElement(programme, "channel").text = str(row["channel"]).strip()
            # ET.SubElement(programme, "title").text = str(row["title"]).strip()
                # Only include title if it's present and not empty
            if pd.notna(row["title"]) and str(row["title"]).strip():
              ET.SubElement(programme, "title").text = str(row["title"]).strip()
            
            # Only add description if it exists
            if pd.notna(row["description"]) and str(row["description"]).strip():
                ET.SubElement(programme, "desc").text = str(row["description"]).strip()
        
            # Only add icon if URL exists and is not empty
            if pd.notna(row["icon_url"]) and str(row["icon_url"]).strip():
                 ET.SubElement(programme, "icon", src=str(row["icon_url"]).strip()) 
                # ET.SubElement(programme, "icon").text = str(row["icon_url"]).strip() # provided URL without an attribute

        
        if self.output_xml.get():
            xml_file = self.output_xml.get()
        else:
            xml_file = excel_file.replace('.xlsx', '.xml')
        
        xmlstr = minidom.parseString(ET.tostring(tv)).toprettyxml(indent="  ")
        
        with open(xml_file, "w", encoding="utf-8") as f:
            f.write(xmlstr)
        
        return xml_file
    
    def parse_time(self, timestr):
        try:
            return datetime.strptime(timestr.strip().upper(), "%I:%M%p").time()
        except ValueError:
            # Try alternative formats if needed
            return datetime.strptime(timestr.strip().upper(), "%H:%M").time()
    
    def combine_datetime(self, date_str, time_str):
        dt_str = f"{date_str} {time_str}"
        try:
            dt = pd.to_datetime(dt_str)
            return dt.strftime("%Y%m%d%H%M%S")
        except:
            return ""

if __name__ == "__main__":
    root = tk.Tk()
    app = EPGGeneratorApp(root)
    root.mainloop()